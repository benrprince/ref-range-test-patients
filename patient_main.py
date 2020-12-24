# Author: Ben Prince
# Version: 1.2
# Description: Used to figure out the number of test patients needed
#              based on the sex and age ranges given in the DCW

import patient_sort as ps
import xlwt
import openpyxl
# TODO: Update xlwt to use openpyxl....currently works, but pulling in redundant library

def parse_patients(filename, overlap):
    """Used to separate the excel rows into 3 lists containing
    each sex category. Skips first excel row because it is
    assumed to be a header row. Returns: male, female, and
    undifferentiated lists"""

    m_list = []
    f_list = []
    u_list = []

    # open excel file and get the first sheet
    xl_workbook = openpyxl.open(filename)
    sheet = xl_workbook.worksheets[0]

    # iterate through the rows and sort the data into the three lists
    for i in range(2, sheet.max_row):
        temp_list = []
        if str(sheet.cell(i, 1).value) == 'Male':
            temp_list.append(str(sheet.cell(i, 1).value))
            temp_list.append(int(sheet.cell(i, 2).value))
            temp_list.append(int(sheet.cell(i, 3).value))
            m_list.append(temp_list)
        elif str(sheet.cell(i, 1).value) == 'Female':
            temp_list.append(str(sheet.cell(i, 1).value))
            temp_list.append(int(sheet.cell(i, 2).value))
            temp_list.append(int(sheet.cell(i, 3).value))
            f_list.append(temp_list)
        else:
            temp_list.append(str(sheet.cell(i, 1).value))
            temp_list.append(int(sheet.cell(i, 2).value))
            temp_list.append(int(sheet.cell(i, 3).value))
            u_list.append(temp_list)

    return m_list, f_list, u_list


def get_patients_wb(filename):
    """Runs the algorithm on the patient_sort file and
    arranges the data into a workbook. Returns: the
    formatted workbook with test patients"""
    
    # minutes for 1 week. This can change
    overlap = 10080

    # use above function to split out the age and sex lines
    m_list, f_list, u_list = parse_patients(filename, overlap)

    # Write to a new workbook
    wb = xlwt.Workbook()
    patients = wb.add_sheet('Patients')

    # Set up Doc
    patients.write(0, 0, 'Sex', xlwt.Style.easyxf("font: bold on"))
    patients.write(0, 1, 'Age', xlwt.Style.easyxf("font: bold on"))

    # Get test patient data into lists from patient_sort file
    m_list = ps.test_patient_list(m_list, overlap)
    m_len = len(m_list)
    f_list = ps.test_patient_list(f_list, overlap)
    f_len = len(f_list)
    u_list = ps.test_patient_list(u_list, overlap)
    u_len = len(u_list)

    # import male data into return doc
    for i in range(1, m_len):
        patients.write(i, 0, m_list[i-1][0])
        patients.write(i, 1, m_list[i-1][1])

    # import female data into return doc
    for i in range(m_len, f_len + m_len):
        patients.write(i, 0, f_list[i - m_len][0])
        patients.write(i, 1, f_list[i - m_len][1])

    # import undefined or unknown data into return doc
    for i in range(m_len + f_len, m_len + f_len + u_len):
        patients.write(i, 0, u_list[i - (m_len+f_len)][0])
        patients.write(i, 1, u_list[i - (m_len+f_len)][1])

    return wb
